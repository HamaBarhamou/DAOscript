import os
import pandas as pd
from weasyprint import HTML
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook

def convertir_excel_en_pdf(path_excel, dossier_sortie="exports_pdf"):
    if not os.path.exists(dossier_sortie):
        os.makedirs(dossier_sortie)

    wb = load_workbook(filename=path_excel, data_only=True)
    sheetnames = wb.sheetnames

    for nom_feuille in sheetnames:
        print(f"📄 Lecture de la feuille : {nom_feuille}")
        
        # Lire sans header pour éviter les Unnamed: X
        df = pd.read_excel(path_excel, sheet_name=nom_feuille, header=None, dtype=str, engine='openpyxl')

        # Remplacer les NaN par ""
        df = df.fillna("")

        # Générer HTML avec style propre
        style_css = """
        @page {
            size: A4;
            margin: 1cm;
        }
        body {
            font-family: 'Noto Sans', sans-serif;
            font-size: 9pt;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
            word-wrap: break-word;
        }
        th, td {
            border: 1px solid #444;
            padding: 4px;
            text-align: center;
            font-size: 8pt;
        }
        thead {
            background-color: #f2f2f2;
            font-weight: bold;
        }
        """

        html_content = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <style>{style_css}</style>
        </head>
        <body>
            <h2>Feuille : {nom_feuille}</h2>
            {df.to_html(index=False, header=False, escape=False)}
        </body>
        </html>
        """

        with NamedTemporaryFile(delete=False, suffix=".html") as tmp_html:
            tmp_html.write(html_content.encode("utf-8"))
            tmp_html_path = tmp_html.name

        nom_pdf = f"{nom_feuille}.pdf".replace(" ", "_")
        chemin_pdf = os.path.join(dossier_sortie, nom_pdf)
        HTML(tmp_html_path).write_pdf(chemin_pdf)

        print(f"✅ PDF généré : {chemin_pdf}")

    print(f"\n🎉 Tous les fichiers PDF sont dans : {dossier_sortie}")


# === Programme principal ===
if __name__ == "__main__":
    chemin = input("Entrez le chemin du fichier Excel (.xlsx) : ").strip()
    try:
        convertir_excel_en_pdf(chemin)
    except Exception as e:
        print(f"❌ Erreur : {e}")

